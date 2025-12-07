import os
import openpyxl
from openpyxl import Workbook

EXCEL = "users.xlsx"

def init_excel():
    if not os.path.exists(EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.append(["username", "password", "highscore"])
        wb.save(EXCEL)

def create_user(username, password):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return {"success": False, "message": "Username exists"}
    ws.append([username, password, 0])
    wb.save(EXCEL)
    return {"success": True, "message": "Registered successfully"}

def verify_user(username, password):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

def get_user_highscore(username):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return row[2] or 0
    return 0

def update_user_score(username, score):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == username:
            prev = row[2].value or 0
            if score > prev:
                row[2].value = score
            break
    wb.save(EXCEL)

def get_all_users():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    users = [
        {"username": row[0].value, "score": row[2].value or 0}
        for row in ws.iter_rows(min_row=2)
    ]
    return users

def record_notes_round(username, round_scores):
    """
    Stores scores of up to 3 rounds for Notes Mania.
    This adds columns Round1, Round2, Round3 in Excel.
    """
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if "Round1" not in headers:
        ws.cell(row=1, column=len(headers)+1, value="Round1")
        ws.cell(row=1, column=len(headers)+2, value="Round2")
        ws.cell(row=1, column=len(headers)+3, value="Round3")
        headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == username:
            for i, score in enumerate(round_scores):
                idx = headers.index(f"Round{i+1}") + 1
                row[idx-1].value = score
            break
    wb.save(EXCEL)
